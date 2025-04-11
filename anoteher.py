import os
import io
import math
import time
import requests
from PyPDF2 import PdfMerger
from compdfkit.client import CPDFClient
from compdfkit.enums import CPDFConversionEnum
from compdfkit.param import CPDFToExcelParameter
from compdfkit.constant import CPDFConstant
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
import openpyxl
from openpyxl.styles import Border, Side

# ==== CONFIGURATION ====
FOLDER_ID = '1HwvyXmGAqsJBeFxMkZ2ZoElgDSAOkAd1'
PUBLIC_KEY =  'public_key_e248af9f4257f5a045bf64fd1e5ab5be'
SECRET_KEY = 'secret_key_33c1f0cef5c2bb0260deccbe6790669d'
SERVICE_ACCOUNT_FILE = 'auth.json'
SCOPES = ['https://www.googleapis.com/auth/drive']

# ==== Google Drive Setup ====
credentials = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)
drive_service = build('drive', 'v3', credentials=credentials)

# ==== CPDFKit Client Setup ====
client = CPDFClient(PUBLIC_KEY, SECRET_KEY)

# ==== List all PDFs in folder ====
query = f"'{FOLDER_ID}' in parents and mimeType='application/pdf'"
results = drive_service.files().list(q=query, fields="files(id, name)").execute()
pdf_files = results.get('files', [])

# ==== Group PDFs into chunks of 15 ====
chunk_size = 15
pdf_chunks = [pdf_files[i:i + chunk_size] for i in range(0, len(pdf_files), chunk_size)]

# ==== Process Each Chunk ====
for idx, chunk in enumerate(pdf_chunks):
    merged_pdf_path = f"/tmp/merged_{idx+1}.pdf"
    merger = PdfMerger()

    print(f"🔧 Merging chunk {idx+1} with {len(chunk)} PDFs...")

    for pdf in chunk:
        file_id = pdf['id']
        file_name = pdf['name']
        temp_path = f"/tmp/{file_name}"

        request = drive_service.files().get_media(fileId=file_id)
        fh = io.FileIO(temp_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)

        done = False
        while not done:
            status, done = downloader.next_chunk()

        merger.append(temp_path)

    merger.write(merged_pdf_path)
    merger.close()
    print(f"✅ Merged file saved to {merged_pdf_path}")

    try:
        create_task_result = client.create_task(CPDFConversionEnum.PDF_TO_EXCEL)
        task_id = create_task_result.task_id

        file_param = CPDFToExcelParameter()
        file_param.content_options = "2"
        file_param.is_contain_img = "0"
        file_param.is_contain_annot = "1"

        upload = client.upload_file(merged_pdf_path, task_id, "", file_param)
        client.execute_task(task_id)

        print(f"🔍 Task created and executing for merged PDF: {merged_pdf_path}")

        attempts = 0
        task_status = None
        while attempts < 5 and task_status != CPDFConstant.TASK_FINISH:
            task_info = client.get_task_info(task_id)
            task_status = task_info.task_status
            print(f"⏳ Task status: {task_status}")
            attempts += 1
            if task_status != CPDFConstant.TASK_FINISH:
                print(f"❌ Task still processing, retrying in 5 seconds...")
                time.sleep(5)

        if task_status == CPDFConstant.TASK_FINISH:
            excel_info = client.get_file_info(upload.file_key)
            download_url = excel_info._download_url

            if download_url:
                excel_path = merged_pdf_path.replace(".pdf", ".xlsx")
                r = requests.get(download_url)
                with open(excel_path, "wb") as f:
                    f.write(r.content)
                print(f"📥 Downloaded Excel: {excel_path}")

                wb = openpyxl.load_workbook(excel_path)

                # Remove sheets not containing 'Consignee'
                for sheet in wb.sheetnames[:]:
                    ws = wb[sheet]
                    if not any("Consignee" in str(cell.value) for row in ws.iter_rows() for cell in row if cell.value):
                        wb.remove(ws)

                # Remove bold text (excluding digits and .) from 2nd sheet onwards
                for i, sheet in enumerate(wb.sheetnames):
                    if i == 0:
                        continue
                    ws = wb[sheet]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.font.bold and not (str(cell.value).isdigit() or str(cell.value).replace(".", "").isdigit()):
                                cell.value = ""

                # Append all other sheets to the first sheet
                main_sheet = wb[wb.sheetnames[0]]
                for sheet_name in reversed(wb.sheetnames[1:]):
                    ws = wb[sheet_name]
                    for row in ws.iter_rows(values_only=True):
                        if any(row):
                            main_sheet.append([str(cell) if cell is not None else "" for cell in row])
                    wb.remove(ws)

                # Remove empty rows and add borders
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                # Set column width for column F
                main_sheet.column_dimensions['F'].width = 15

                data = list(main_sheet.values)
                main_sheet.delete_rows(1, main_sheet.max_row)
                for row in data:
                    if any(str(cell).strip() for cell in row):
                        main_sheet.append(row)

                for row in main_sheet.iter_rows(min_row=1, max_row=main_sheet.max_row, min_col=1, max_col=7):
                    for cell in row:
                        cell.border = thin_border

                # Remove columns E, H, I
                for col in ['E', 'H', 'I']:
                    col_index = openpyxl.utils.column_index_from_string(col)
                    main_sheet.delete_cols(col_index)

                # Modify Column B text to keep content after the second occurrence of '****'
                for row in main_sheet.iter_rows(min_col=2, max_col=2):
                    for cell in row:
                        if cell.value:
                            value = str(cell.value)
                            if value.count('****') >= 2:
                                value = value.split('****', 2)[2]
                                cell.value = value

                # Extract Column D data from D2 to end as strings
                extracted_texts = []
                for row in main_sheet.iter_rows(min_row=2, min_col=4, max_col=4):
                    for cell in row:
                        if cell.value:
                            extracted_texts.append(f'text: "{str(cell.value).strip()}"')

                print("\n\n==== Extracted Texts for External Script ====")
                for t in extracted_texts:
                    print(t)
                print("===========================================\n\n")

                wb.save(excel_path)

                file_metadata = {
                    'name': os.path.basename(excel_path),
                    'parents': [FOLDER_ID]
                }
                media = MediaFileUpload(
                    excel_path,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                drive_service.files().create(body=file_metadata,
                                             media_body=media,
                                             fields='id').execute()
                print(f"📤 Uploaded Excel to Drive: {os.path.basename(excel_path)}")
            else:
                print("❌ No download URL returned from CompdfKit.")
        else:
            print(f"❌ Conversion failed for merged chunk {idx+1} after {attempts} attempts")
    except Exception as e:
        print(f"❌ Error occurred while processing merged chunk {idx+1}: {str(e)}")
